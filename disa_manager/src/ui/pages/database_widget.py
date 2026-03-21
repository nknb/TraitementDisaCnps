from __future__ import annotations

from typing import List, Optional, Dict

from PySide6.QtCore import Qt
from PySide6.QtGui import QIcon, QColor, QFont
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QComboBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QDialog,
    QFormLayout,
    QDialogButtonBox,
    QMessageBox,
    QFrame,
    QFileDialog,
    QSizePolicy,
    QDateEdit,
)
from PySide6.QtCore import QDate

from db.connection import get_connection
from core.events import get_data_bus
from ui.dashboard_theme import (
    BTN_PRIMARY as _STYLE_BTN_PRIMARY,
    BTN_SUCCESS as _STYLE_BTN_SUCCESS,
    BTN_DANGER as _STYLE_BTN_DANGER,
    BTN_NEUTRAL as _STYLE_BTN_NEUTRAL,
    INPUT_STYLE as _INPUT_STYLE,
)

# ---------------------------------------------------------------------------
# Styles réutilisables (complémentaires non présents dans dashboard_theme)
# ---------------------------------------------------------------------------
_STYLE_BTN_EXPORT = (
    "QPushButton { background-color: #0369a1; color: white; border-radius: 5px; "
    "padding: 6px 14px; font-weight: 600; font-size: 12px; }"
    "QPushButton:hover { background-color: #0284c7; }"
    "QPushButton:pressed { background-color: #075985; }"
)
_STYLE_BTN_SUSPEND = (
    "QPushButton { background-color: #92400e; color: white; border-radius: 5px; "
    "padding: 6px 14px; font-weight: 600; font-size: 12px; }"
    "QPushButton:hover { background-color: #b45309; }"
    "QPushButton:pressed { background-color: #78350f; }"
)
_STYLE_BTN_REACTIVATE = (
    "QPushButton { background-color: #15803d; color: white; border-radius: 5px; "
    "padding: 6px 14px; font-weight: 600; font-size: 12px; }"
    "QPushButton:hover { background-color: #16a34a; }"
    "QPushButton:pressed { background-color: #14532d; }"
)
_STYLE_INPUT = (
    "QLineEdit { border: 1px solid #d1d5db; border-radius: 4px; padding: 5px 8px; "
    "font-size: 12px; background: white; color: #1f2937; }"
    "QLineEdit:focus { border-color: #1e3a5f; border-width: 2px; }"
    "QLineEdit:placeholder { color: #9ca3af; }"
)
_STYLE_COMBO = (
    "QComboBox { border: 1px solid #d1d5db; border-radius: 4px; padding: 5px 24px 5px 8px; "
    "font-size: 12px; background: white; color: #1f2937; }"
    "QComboBox:focus { border-color: #1e3a5f; border-width: 2px; }"
    "QComboBox:hover { border-color: #93c5fd; }"
    "QComboBox::drop-down { subcontrol-origin: padding; subcontrol-position: right center; "
    "width: 20px; border: none; }"
    "QComboBox::down-arrow { width: 10px; height: 10px; }"
    "QComboBox QAbstractItemView { border: 1px solid #d1d5db; border-radius: 4px; "
    "background: white; color: #1f2937; outline: none; "
    "selection-background-color: #dbeafe; selection-color: #1e3a5f; }"
    "QComboBox QAbstractItemView::item { padding: 5px 10px; min-height: 28px; }"
    "QComboBox QAbstractItemView::item:hover { background: #eff6ff; color: #1e3a5f; }"
)
_STYLE_TABLE = (
    "QTableWidget { border: 1px solid #e2e8f0; gridline-color: #f1f5f9; font-size: 12px; }"
    "QTableWidget::item { padding: 5px 8px; }"
    "QTableWidget::item:selected { background-color: #dbeafe; color: #1e3a5f; }"
    "QHeaderView::section { background-color: #1e3a5f; color: white; font-weight: 700; "
    "padding: 7px 8px; border: none; border-right: 1px solid #2a4f80; }"
    "QTableWidget::item:alternate { background-color: #f8fafc; }"
)


def _make_btn(label: str, style: str, icon_path: str = "") -> QPushButton:
    btn = QPushButton(label)
    btn.setStyleSheet(style)
    btn.setCursor(Qt.CursorShape.PointingHandCursor)
    if icon_path:
        btn.setIcon(QIcon(icon_path))
    return btn


def _make_separator() -> QFrame:
    sep = QFrame()
    sep.setFrameShape(QFrame.Shape.VLine)
    sep.setFrameShadow(QFrame.Shadow.Sunken)
    sep.setStyleSheet("color: #e2e8f0;")
    return sep


# ---------------------------------------------------------------------------
# Formulaire d'ajout / modification
# ---------------------------------------------------------------------------

class EmployeurFormDialog(QDialog):
    """Dialogue générique pour ajouter / modifier un employeur."""

    def __init__(
        self,
        parent: QWidget,
        columns: List[str],
        data: Optional[Dict[str, Optional[str]]] = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Employeur — Base de données")
        self.setMinimumWidth(420)
        self._columns = [c for c in columns if c != "id"]
        self._editors: Dict[str, QLineEdit] = {}

        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 16)

        # Titre interne
        header = QLabel("Informations employeur")
        header.setStyleSheet(
            "font-size: 14px; font-weight: 700; color: #1e3a5f; padding-bottom: 4px;"
        )
        layout.addWidget(header)

        form = QFormLayout()
        form.setSpacing(8)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        for col in self._columns:
            editor = QLineEdit(self)
            editor.setStyleSheet(_STYLE_INPUT)
            if data and col in data and data[col] is not None:
                editor.setText(str(data[col]))
            label = QLabel(col.replace("_", " ").title())
            label.setStyleSheet("font-size: 12px; color: #374151; font-weight: 600;")
            form.addRow(label, editor)
            self._editors[col] = editor

        layout.addLayout(form)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel,
            parent=self,
        )
        buttons.button(QDialogButtonBox.StandardButton.Ok).setText("Enregistrer")
        buttons.button(QDialogButtonBox.StandardButton.Ok).setStyleSheet(_STYLE_BTN_PRIMARY)
        buttons.button(QDialogButtonBox.StandardButton.Cancel).setStyleSheet(_STYLE_BTN_NEUTRAL)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def get_values(self) -> Dict[str, Optional[str]]:
        values: Dict[str, Optional[str]] = {}
        for col, editor in self._editors.items():
            text = editor.text().strip()
            values[col] = text if text != "" else None
        return values


# ---------------------------------------------------------------------------
# Widget principal
# ---------------------------------------------------------------------------

class EmployersDatabaseWidget(QWidget):
    """Onglet "Base de données" — Vue jointe Employeurs + Traitement DISA.

    Fonctionnalités :
    - Affichage paginé de la vue jointe
    - Filtres (recherche texte, localité, exercice, statut, plage de dates)
    - CRUD (Ajouter, Modifier, Supprimer)
    - Export Excel des données filtrées (toutes les pages)
    """

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)

        self._table_name: str = "join_employeur_traitement"
        self._columns: List[str] = []
        self._id_index: int = -1
        self._page_size: int = 50
        self._current_page: int = 1
        self._total_rows: int = 0
        self._needs_refresh: bool = False

        self._build_ui()
        self._init_table_combo()
        self._load_structure()
        self._load_filters()
        self._refresh_table()

        get_data_bus().data_changed.connect(self._refresh_table)
        self.table.itemSelectionChanged.connect(self._update_suspend_button)
        self.table.itemChanged.connect(self._on_checkbox_changed)
        self.table.horizontalHeader().sectionClicked.connect(self._on_header_section_clicked)

    def showEvent(self, event) -> None:  # type: ignore[override]
        """Rafraîchit les données si une mise à jour a eu lieu pendant que le widget était caché."""
        super().showEvent(event)
        if self._needs_refresh:
            self._refresh_table()

    # ------------------------------------------------------------------
    # Construction de l'UI
    # ------------------------------------------------------------------

    def _build_ui(self) -> None:
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 12)
        main_layout.setSpacing(10)

        # ── En-tête ────────────────────────────────────────────────────
        header_frame = QFrame()
        header_frame.setStyleSheet(
            "background: qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #1e3a5f, stop:1 #2a4f80);"
            "border-radius: 8px; padding: 10px 16px;"
        )
        header_layout = QHBoxLayout(header_frame)
        header_layout.setContentsMargins(16, 10, 16, 10)

        title_lbl = QLabel("Base de données — Employeurs & DISA")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_lbl.setFont(title_font)
        title_lbl.setStyleSheet("color: white; background: transparent;")
        header_layout.addWidget(title_lbl)
        header_layout.addStretch(1)

        # Compteurs de statut dans l'en-tête
        self.status_non_traite_lbl = self._make_badge("Non traités : 0", "#dc2626")
        self.status_valide_lbl = self._make_badge("Validés : 0", "#15803d")
        self.status_rejet_lbl = self._make_badge("Avec rejets : 0", "#d97706")
        header_layout.addWidget(self.status_non_traite_lbl)
        header_layout.addWidget(self.status_valide_lbl)
        header_layout.addWidget(self.status_rejet_lbl)

        main_layout.addWidget(header_frame)

        # ── Sélecteur de table ─────────────────────────────────────────
        table_row = QHBoxLayout()
        table_label = QLabel("Affichage :")
        table_label.setStyleSheet("font-size: 12px; font-weight: 600; color: #374151;")
        self.table_combo = QComboBox()
        self.table_combo.setStyleSheet(_STYLE_COMBO)
        self.table_combo.setFixedHeight(30)
        self.table_combo.currentIndexChanged.connect(self._on_table_changed)
        table_row.addWidget(table_label)
        table_row.addWidget(self.table_combo, 1)
        table_row.addStretch(3)
        main_layout.addLayout(table_row)

        # ── Filtres ────────────────────────────────────────────────────
        filters_frame = QFrame()
        filters_frame.setStyleSheet(
            "QFrame { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; }"
        )
        filters_outer = QVBoxLayout(filters_frame)
        filters_outer.setContentsMargins(14, 10, 14, 10)
        filters_outer.setSpacing(8)

        # Barre titre des filtres
        filters_title_row = QHBoxLayout()
        filters_title_lbl = QLabel("FILTRES")
        filters_title_lbl.setStyleSheet(
            "font-size: 11px; font-weight: 800; color: #374151; "
            "letter-spacing: 1px; background: transparent;"
        )
        self._filter_count_badge = QLabel()
        self._filter_count_badge.setStyleSheet(
            "background-color: #1e3a5f; color: white; border-radius: 8px; "
            "padding: 1px 8px; font-size: 10px; font-weight: 700;"
        )
        self._filter_count_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._filter_count_badge.hide()

        reset_btn = _make_btn("↺  Réinitialiser", _STYLE_BTN_NEUTRAL)
        reset_btn.setFixedHeight(26)
        reset_btn.setMaximumWidth(120)
        reset_btn.clicked.connect(self._reset_filters)

        filters_title_row.addWidget(filters_title_lbl)
        filters_title_row.addWidget(self._filter_count_badge)
        filters_title_row.addStretch(1)
        filters_title_row.addWidget(reset_btn)
        filters_outer.addLayout(filters_title_row)

        # Séparateur
        sep_line = QFrame()
        sep_line.setFrameShape(QFrame.Shape.HLine)
        sep_line.setStyleSheet("color: #e2e8f0;")
        filters_outer.addWidget(sep_line)

        # ── Ligne 1 : Recherche ─────────────────────────────────────────
        def _field_col(label_text: str, widget: QWidget) -> QVBoxLayout:
            """Empile un label + un champ de saisie verticalement."""
            col = QVBoxLayout()
            col.setSpacing(3)
            lbl = QLabel(label_text)
            lbl.setStyleSheet(
                "font-size: 10px; font-weight: 700; color: #6b7280; "
                "letter-spacing: 0.5px; background: transparent;"
            )
            col.addWidget(lbl)
            col.addWidget(widget)
            return col

        search_row = QHBoxLayout()
        search_row.setSpacing(8)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Rechercher par n°, raison sociale, localité, exercice…")
        self.search_edit.setStyleSheet(_STYLE_INPUT)
        self.search_edit.setFixedHeight(30)
        self.search_edit.textChanged.connect(self._on_filters_changed)

        # Bouton ✕ intégré pour effacer la recherche
        self._search_clear_btn = QPushButton("✕")
        self._search_clear_btn.setFixedSize(30, 30)
        self._search_clear_btn.setStyleSheet(
            "QPushButton { background: #e5e7eb; color: #6b7280; border-radius: 4px; "
            "font-weight: 700; border: none; }"
            "QPushButton:hover { background: #dc2626; color: white; }"
        )
        self._search_clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._search_clear_btn.setToolTip("Effacer la recherche")
        self._search_clear_btn.hide()
        self._search_clear_btn.clicked.connect(self.search_edit.clear)
        self.search_edit.textChanged.connect(
            lambda t: self._search_clear_btn.setVisible(bool(t))
        )

        search_hbox = QHBoxLayout()
        search_hbox.setSpacing(4)
        search_hbox.addWidget(self.search_edit)
        search_hbox.addWidget(self._search_clear_btn)

        search_col = QVBoxLayout()
        search_col.setSpacing(3)
        search_lbl = QLabel("RECHERCHE")
        search_lbl.setStyleSheet(
            "font-size: 10px; font-weight: 700; color: #6b7280; "
            "letter-spacing: 0.5px; background: transparent;"
        )
        search_col.addWidget(search_lbl)
        search_col.addLayout(search_hbox)

        search_row.addLayout(search_col, 1)
        filters_outer.addLayout(search_row)

        # ── Ligne 2 : Combos ────────────────────────────────────────────
        combos_row = QHBoxLayout()
        combos_row.setSpacing(12)

        self.localite_combo = QComboBox()
        self.localite_combo.setStyleSheet(_STYLE_COMBO)
        self.localite_combo.setFixedHeight(30)
        self.localite_combo.currentIndexChanged.connect(self._on_filters_changed)

        self.exercice_combo = QComboBox()
        self.exercice_combo.setStyleSheet(_STYLE_COMBO)
        self.exercice_combo.setFixedHeight(30)
        self.exercice_combo.currentIndexChanged.connect(self._on_filters_changed)

        self.secteur_combo = QComboBox()
        self.secteur_combo.setStyleSheet(_STYLE_COMBO)
        self.secteur_combo.setFixedHeight(30)
        self.secteur_combo.currentIndexChanged.connect(self._on_filters_changed)

        self.suspension_combo = QComboBox()
        self.suspension_combo.setStyleSheet(_STYLE_COMBO)
        self.suspension_combo.setFixedHeight(30)
        self.suspension_combo.addItem("Toutes les entreprises", None)
        self.suspension_combo.addItem("⛔  Suspendues", 1)
        self.suspension_combo.addItem("✔  Actives", 0)
        self.suspension_combo.currentIndexChanged.connect(self._on_filters_changed)

        combos_row.addLayout(_field_col("LOCALITÉ", self.localite_combo), 2)
        combos_row.addLayout(_field_col("EXERCICE", self.exercice_combo), 1)
        combos_row.addLayout(_field_col("STATUT", self.secteur_combo), 2)
        combos_row.addLayout(_field_col("SUSPENSION", self.suspension_combo), 2)

        # ── Sélecteurs de date ──────────────────────────────────────────
        _DATE_SENTINEL = QDate(2000, 1, 1)
        _STYLE_DATE = (
            "QDateEdit { border: 1px solid #d1d5db; border-radius: 4px; padding: 5px 8px; "
            "font-size: 12px; background: white; color: #1f2937; }"
            "QDateEdit:focus { border-color: #1e3a5f; border-width: 2px; }"
            "QDateEdit:hover { border-color: #93c5fd; }"
            "QDateEdit::drop-down { subcontrol-origin: padding; subcontrol-position: right center; "
            "width: 20px; border: none; }"
        )

        self.date_from_edit = QDateEdit()
        self.date_from_edit.setDisplayFormat("dd/MM/yyyy")
        self.date_from_edit.setCalendarPopup(True)
        self.date_from_edit.setDate(_DATE_SENTINEL)
        self.date_from_edit.setMinimumDate(QDate(2000, 1, 1))
        self.date_from_edit.setSpecialValueText("(toutes dates)")
        self.date_from_edit.setStyleSheet(_STYLE_DATE)
        self.date_from_edit.setFixedHeight(30)
        self.date_from_edit.dateChanged.connect(self._on_filters_changed)

        self.date_to_edit = QDateEdit()
        self.date_to_edit.setDisplayFormat("dd/MM/yyyy")
        self.date_to_edit.setCalendarPopup(True)
        self.date_to_edit.setDate(_DATE_SENTINEL)
        self.date_to_edit.setMinimumDate(QDate(2000, 1, 1))
        self.date_to_edit.setSpecialValueText("(toutes dates)")
        self.date_to_edit.setStyleSheet(_STYLE_DATE)
        self.date_to_edit.setFixedHeight(30)
        self.date_to_edit.dateChanged.connect(self._on_filters_changed)

        # Stocker la sentinelle pour les comparaisons
        self._date_sentinel = _DATE_SENTINEL

        dates_hbox = QHBoxLayout()
        dates_hbox.setSpacing(6)
        sep_arrow = QLabel("→")
        sep_arrow.setStyleSheet("color: #9ca3af; font-weight: 700; background: transparent;")
        sep_arrow.setAlignment(Qt.AlignmentFlag.AlignCenter)
        dates_hbox.addWidget(self.date_from_edit)
        dates_hbox.addWidget(sep_arrow)
        dates_hbox.addWidget(self.date_to_edit)

        dates_col = QVBoxLayout()
        dates_col.setSpacing(3)
        dates_lbl = QLabel("DATE DE RÉCEPTION")
        dates_lbl.setStyleSheet(
            "font-size: 10px; font-weight: 700; color: #6b7280; "
            "letter-spacing: 0.5px; background: transparent;"
        )
        dates_col.addWidget(dates_lbl)
        dates_col.addLayout(dates_hbox)

        combos_row.addLayout(dates_col, 2)
        filters_outer.addLayout(combos_row)

        # ── Puces "filtres actifs" ──────────────────────────────────────
        self._active_filters_row = QHBoxLayout()
        self._active_filters_row.setSpacing(6)
        self._active_filters_container = QFrame()
        self._active_filters_container.setLayout(self._active_filters_row)
        self._active_filters_container.setStyleSheet("background: transparent; border: none;")
        self._active_filters_container.hide()
        filters_outer.addWidget(self._active_filters_container)

        main_layout.addWidget(filters_frame)

        # ── Barre d'actions ────────────────────────────────────────────
        actions_row = QHBoxLayout()
        actions_row.setSpacing(8)

        self.add_btn = _make_btn("＋  Ajouter", _STYLE_BTN_PRIMARY, ":/icon/icon/product-32.ico")
        self.add_btn.setFixedHeight(32)
        self.add_btn.clicked.connect(self._on_add_clicked)
        actions_row.addWidget(self.add_btn)

        self.edit_btn = _make_btn("✎  Modifier", _STYLE_BTN_SUCCESS, ":/icon/icon/activity-feed-32.ico")
        self.edit_btn.setFixedHeight(32)
        self.edit_btn.clicked.connect(self._on_edit_clicked)
        actions_row.addWidget(self.edit_btn)

        self.delete_btn = _make_btn("✕  Supprimer", _STYLE_BTN_DANGER, ":/icon/icon/close-window-64.ico")
        self.delete_btn.setFixedHeight(32)
        self.delete_btn.setEnabled(False)  # Activé uniquement quand des cases sont cochées
        self.delete_btn.clicked.connect(self._on_delete_clicked)
        actions_row.addWidget(self.delete_btn)

        self.suspend_btn = _make_btn("⛔  Suspendre", _STYLE_BTN_SUSPEND)
        self.suspend_btn.setFixedHeight(32)
        self.suspend_btn.setToolTip("Suspendre ou réactiver l'entreprise sélectionnée")
        self.suspend_btn.clicked.connect(self._on_toggle_suspension_clicked)
        actions_row.addWidget(self.suspend_btn)

        actions_row.addWidget(_make_separator())

        self.refresh_btn = _make_btn("↻  Actualiser", _STYLE_BTN_NEUTRAL, ":/icon/icon/dashboard-5-32.ico")
        self.refresh_btn.setFixedHeight(32)
        self.refresh_btn.clicked.connect(self._refresh_table)
        actions_row.addWidget(self.refresh_btn)

        actions_row.addWidget(_make_separator())

        self.export_btn = _make_btn("⬇  Exporter Excel", _STYLE_BTN_EXPORT)
        self.export_btn.setFixedHeight(32)
        self.export_btn.setToolTip("Exporte toutes les lignes correspondant aux filtres actifs vers un fichier Excel")
        self.export_btn.clicked.connect(self._on_export_excel_clicked)
        actions_row.addWidget(self.export_btn)

        actions_row.addStretch(1)

        main_layout.addLayout(actions_row)

        # ── Tableau ────────────────────────────────────────────────────
        self.table = QTableWidget(self)
        self.table.setSortingEnabled(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(_STYLE_TABLE)
        self.table.verticalHeader().setDefaultSectionSize(30)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setHighlightSections(False)
        self.table.setShowGrid(True)
        main_layout.addWidget(self.table, 1)

        # ── Pied de page : pagination ──────────────────────────────────
        footer_layout = QHBoxLayout()
        footer_layout.setSpacing(8)

        self.prev_page_btn = _make_btn("‹  Précédent", _STYLE_BTN_NEUTRAL)
        self.prev_page_btn.setFixedHeight(28)
        self.next_page_btn = _make_btn("Suivant  ›", _STYLE_BTN_NEUTRAL)
        self.next_page_btn.setFixedHeight(28)
        self.prev_page_btn.clicked.connect(self._on_prev_page)
        self.next_page_btn.clicked.connect(self._on_next_page)

        size_label = QLabel("Lignes / page :")
        size_label.setStyleSheet("font-size: 12px; color: #374151;")
        self.page_size_combo = QComboBox()
        self.page_size_combo.setStyleSheet(_STYLE_COMBO)
        self.page_size_combo.setFixedHeight(28)
        self.page_size_combo.setMaximumWidth(70)
        self.page_size_combo.addItems(["10", "25", "50", "100"])
        self.page_size_combo.setCurrentText("50")
        self.page_size_combo.currentIndexChanged.connect(self._on_page_size_changed)

        self.page_label = QLabel("Page 1 / 1")
        self.page_label.setStyleSheet("font-size: 12px; color: #374151; font-weight: 600;")

        footer_layout.addWidget(self.prev_page_btn)
        footer_layout.addWidget(self.next_page_btn)
        footer_layout.addWidget(size_label)
        footer_layout.addWidget(self.page_size_combo)
        footer_layout.addStretch(1)
        footer_layout.addWidget(self.page_label)

        main_layout.addLayout(footer_layout)

    # ------------------------------------------------------------------
    # Helpers visuels
    # ------------------------------------------------------------------

    @staticmethod
    def _make_badge(text: str, color: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(
            f"background-color: {color}; color: white; border-radius: 10px; "
            "padding: 3px 10px; font-size: 11px; font-weight: 700;"
        )
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        return lbl

    def _reset_filters(self) -> None:
        """Remet tous les filtres à leur valeur par défaut."""
        for w in (
            self.search_edit, self.date_from_edit, self.date_to_edit,
            self.localite_combo, self.exercice_combo, self.secteur_combo,
            self.suspension_combo,
        ):
            w.blockSignals(True)

        self.search_edit.clear()
        self.date_from_edit.setDate(self._date_sentinel)
        self.date_to_edit.setDate(self._date_sentinel)
        self.localite_combo.setCurrentIndex(0)
        self.exercice_combo.setCurrentIndex(0)
        self.secteur_combo.setCurrentIndex(0)
        self.suspension_combo.setCurrentIndex(0)

        for w in (
            self.search_edit, self.date_from_edit, self.date_to_edit,
            self.localite_combo, self.exercice_combo, self.secteur_combo,
            self.suspension_combo,
        ):
            w.blockSignals(False)

        self._current_page = 1
        self._update_filter_indicator()
        self._refresh_table()

    def _update_filter_indicator(self) -> None:
        """Met à jour le badge compteur et les puces de filtres actifs."""
        active: list[tuple[str, str]] = []  # (label, valeur affichée)

        if self.search_edit.text().strip():
            active.append(("Recherche", self.search_edit.text().strip()))
        if self.localite_combo.currentData() is not None:
            active.append(("Localité", self.localite_combo.currentText()))
        if self.exercice_combo.currentData() is not None:
            active.append(("Exercice", self.exercice_combo.currentText()))
        if self.secteur_combo.currentData() is not None:
            active.append(("Statut", self.secteur_combo.currentText()))
        if self.suspension_combo.currentData() is not None:
            active.append(("Suspension", self.suspension_combo.currentText()))

        date_from = self.date_from_edit.date()
        date_to = self.date_to_edit.date()
        if date_from != self._date_sentinel or date_to != self._date_sentinel:
            from_str = date_from.toString("dd/MM/yyyy") if date_from != self._date_sentinel else "…"
            to_str = date_to.toString("dd/MM/yyyy") if date_to != self._date_sentinel else "…"
            active.append(("Date", f"{from_str} → {to_str}"))

        # Badge dans la barre titre
        n = len(active)
        if n:
            self._filter_count_badge.setText(f"{n} filtre{'s' if n > 1 else ''} actif{'s' if n > 1 else ''}")
            self._filter_count_badge.show()
        else:
            self._filter_count_badge.hide()

        # Puces
        # Vider les puces existantes
        while self._active_filters_row.count():
            item = self._active_filters_row.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        if active:
            intro = QLabel("Filtres actifs :")
            intro.setStyleSheet("font-size: 10px; color: #6b7280; background: transparent;")
            self._active_filters_row.addWidget(intro)
            for label, value in active:
                chip = QLabel(f"  {label}: {value}  ")
                chip.setStyleSheet(
                    "background-color: #dbeafe; color: #1e3a5f; border-radius: 10px; "
                    "font-size: 10px; font-weight: 700; padding: 2px 6px;"
                )
                self._active_filters_row.addWidget(chip)
            self._active_filters_row.addStretch(1)
            self._active_filters_container.show()
        else:
            self._active_filters_container.hide()

    # ------------------------------------------------------------------
    # Initialisation des tables et filtres
    # ------------------------------------------------------------------

    def _init_table_combo(self) -> None:
        self.table_combo.blockSignals(True)
        self.table_combo.clear()
        self.table_combo.addItem("Vue Employeurs + DISA", "join_employeur_traitement")
        self.table_combo.setCurrentIndex(0)
        self.table_combo.blockSignals(False)

    def _on_filters_changed(self) -> None:
        self._current_page = 1
        self._update_filter_indicator()
        self._refresh_table()

    def _on_table_changed(self) -> None:
        table = self.table_combo.currentData()
        if not table:
            return
        self._table_name = str(table)
        self._load_structure()
        self._load_filters()
        self._refresh_table()

    def _load_structure(self) -> None:
        self._id_index = -1

        if self._table_name == "join_employeur_traitement":
            try:
                conn = get_connection()
                with conn:
                    cur = conn.cursor()
                    cur.execute("PRAGMA table_info(identification_employeurs)")
                    emp_rows = cur.fetchall()
                    cur.execute("PRAGMA table_info(traitement_disa)")
                    td_rows = cur.fetchall()
            except Exception as exc:
                QMessageBox.critical(
                    self, "Erreur base de données",
                    f"Impossible de lire la structure des tables :\n{exc}",
                )
                return

            emp_cols = [r[1] for r in emp_rows]
            td_cols = [r[1] for r in td_rows]
            duplicates = set(emp_cols) & set(td_cols)

            columns: list[str] = ["id_employeur"]
            for name in emp_cols:
                if name != "id":
                    columns.append(name)
            columns.append("id_traitement")
            for name in td_cols:
                if name == "id":
                    continue
                alias = "statut" if name == "statut" else (f"{name}_disa" if name in duplicates else name)
                columns.append(alias)

            self._columns = columns
        else:
            try:
                conn = get_connection()
                with conn:
                    cur = conn.cursor()
                    cur.execute(f"PRAGMA table_info({self._table_name})")
                    rows = cur.fetchall()
            except Exception as exc:
                QMessageBox.critical(
                    self, "Erreur base de données",
                    f"Impossible de lire la structure de {self._table_name} :\n{exc}",
                )
                return
            self._columns = [r[1] for r in rows]

        if self._table_name in ("traitement_disa", "join_employeur_traitement"):
            headers = ["☑"] + list(self._columns) + ["État"]
        else:
            headers = ["☑"] + list(self._columns)
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        # Colonne 0 = cases à cocher (largeur fixe)
        self.table.setColumnWidth(0, 36)
        from PySide6.QtWidgets import QHeaderView as _HV
        self.table.horizontalHeader().setSectionResizeMode(0, _HV.ResizeMode.Fixed)

        if "id" in self._columns:
            self._id_index = self._columns.index("id")
            self.table.setColumnHidden(self._id_index + 1, True)  # +1 : offset checkbox

    def _load_filters(self) -> None:
        self.localite_combo.blockSignals(True)
        self.exercice_combo.blockSignals(True)
        self.secteur_combo.blockSignals(True)
        self.localite_combo.clear()
        self.exercice_combo.clear()
        self.secteur_combo.clear()

        if self._table_name == "identification_employeurs":
            self.localite_combo.addItem("Toutes les localités", None)
            self.exercice_combo.addItem("Tous les exercices", None)
            self.secteur_combo.addItem("Tous les secteurs", None)
            try:
                conn = get_connection()
                with conn:
                    cur = conn.cursor()
                    cur.execute(
                        "SELECT DISTINCT localites FROM identification_employeurs "
                        "WHERE localites IS NOT NULL ORDER BY localites"
                    )
                    for (val,) in cur.fetchall():
                        if val and str(val).strip():
                            self.localite_combo.addItem(str(val), str(val))
                    cur.execute(
                        "SELECT DISTINCT exercice FROM identification_employeurs "
                        "WHERE exercice IS NOT NULL ORDER BY exercice"
                    )
                    for (val,) in cur.fetchall():
                        if val and str(val).strip():
                            self.exercice_combo.addItem(str(val), str(val))
                    cur.execute(
                        "SELECT DISTINCT secteur_activite FROM identification_employeurs "
                        "WHERE secteur_activite IS NOT NULL ORDER BY secteur_activite"
                    )
                    for (val,) in cur.fetchall():
                        if val and str(val).strip():
                            self.secteur_combo.addItem(str(val), str(val))
            finally:
                pass
        else:
            self.localite_combo.addItem("Toutes les localités", None)
            self.exercice_combo.addItem("Tous les exercices", None)
            self.secteur_combo.addItem("Tous les statuts", None)
            try:
                conn = get_connection()
                with conn:
                    cur = conn.cursor()

                    # Toutes les localités de la table employeurs (pas seulement
                    # celles liées à un traitement).
                    cur.execute(
                        "SELECT DISTINCT localites "
                        "FROM identification_employeurs "
                        "WHERE localites IS NOT NULL AND TRIM(localites) != '' "
                        "ORDER BY localites"
                    )
                    for (val,) in cur.fetchall():
                        self.localite_combo.addItem(str(val), str(val))

                    # Tous les exercices des deux tables réunis.
                    cur.execute(
                        "SELECT DISTINCT exercice FROM ("
                        "  SELECT exercice FROM identification_employeurs "
                        "  WHERE exercice IS NOT NULL AND TRIM(exercice) != '' "
                        "  UNION "
                        "  SELECT exercice FROM traitement_disa "
                        "  WHERE exercice IS NOT NULL AND TRIM(exercice) != '' "
                        ") ORDER BY exercice"
                    )
                    for (val,) in cur.fetchall():
                        self.exercice_combo.addItem(str(val), str(val))

                    # Tous les statuts réels présents dans traitement_disa
                    # + "NON TRAITÉ" si au moins un employeur n'a pas de traitement.
                    cur.execute(
                        "SELECT DISTINCT statut "
                        "FROM traitement_disa "
                        "WHERE statut IS NOT NULL AND TRIM(statut) != '' "
                        "ORDER BY statut"
                    )
                    statuts = [str(r[0]) for r in cur.fetchall()]

                    # Vérifier s'il existe des employeurs sans aucune ligne de traitement
                    cur.execute(
                        "SELECT COUNT(*) FROM identification_employeurs ie "
                        "WHERE NOT EXISTS ("
                        "  SELECT 1 FROM traitement_disa td WHERE td.employeur_id = ie.id"
                        ")"
                    )
                    has_non_traite = (cur.fetchone()[0] or 0) > 0

                    # "NON TRAITÉ" en premier — seulement s'il n'est pas déjà dans statuts
                    if has_non_traite and "NON TRAITÉ" not in statuts:
                        self.secteur_combo.addItem("NON TRAITÉ", "NON TRAITÉ")
                    for val in statuts:
                        self.secteur_combo.addItem(val, val)
            finally:
                pass

        self.localite_combo.blockSignals(False)
        self.exercice_combo.blockSignals(False)
        self.secteur_combo.blockSignals(False)

    # ------------------------------------------------------------------
    # Construction du SQL filtré
    # ------------------------------------------------------------------

    def _get_date_filter(self) -> tuple[Optional[str], Optional[str]]:
        """Retourne (date_from, date_to) en format ISO, ou None si non défini."""
        d_from = self.date_from_edit.date()
        d_to = self.date_to_edit.date()
        from_str = d_from.toString("yyyy-MM-dd") if d_from != self._date_sentinel else None
        to_str = d_to.toString("yyyy-MM-dd") if d_to != self._date_sentinel else None
        return from_str, to_str

    def _build_filters_sql(self) -> tuple[str, list]:
        clauses: list[str] = []
        params: list = []

        search = self.search_edit.text().strip()
        if search:
            like = f"%{search}%"
            if self._table_name == "identification_employeurs":
                clauses.append(
                    "(numero LIKE ? OR numero_cnps LIKE ? OR raison_sociale LIKE ? OR localites LIKE ?)"
                )
                params.extend([like, like, like, like])
            elif self._table_name == "traitement_disa":
                clauses.append("(exercice LIKE ? OR statut LIKE ? OR observations LIKE ?)")
                params.extend([like, like, like])
            elif self._table_name == "join_employeur_traitement":
                clauses.append(
                    "(ie.numero LIKE ? OR ie.numero_cnps LIKE ? OR ie.raison_sociale LIKE ? "
                    "OR ie.localites LIKE ? OR td.exercice LIKE ? OR td.statut LIKE ? OR td.observations LIKE ?)"
                )
                params.extend([like, like, like, like, like, like, like])

        if self._table_name == "identification_employeurs":
            if (v := self.localite_combo.currentData()) is not None:
                clauses.append("localites = ?"); params.append(v)
            if (v := self.exercice_combo.currentData()) is not None:
                clauses.append("exercice = ?"); params.append(v)
            if (v := self.secteur_combo.currentData()) is not None:
                clauses.append("secteur_activite = ?"); params.append(v)

        elif self._table_name == "traitement_disa":
            if (v := self.localite_combo.currentData()) is not None:
                clauses.append(
                    "employeur_id IN (SELECT id FROM identification_employeurs WHERE localites = ?)"
                ); params.append(v)
            if (v := self.exercice_combo.currentData()) is not None:
                clauses.append("exercice = ?"); params.append(v)
            if (v := self.secteur_combo.currentData()) is not None:
                if str(v).upper() in ("NON TRAITÉ", "NON TRAITE"):
                    clauses.append("(statut IS NULL OR statut = 'NON TRAITÉ')")
                else:
                    clauses.append("statut = ?"); params.append(v)
            date_from, date_to = self._get_date_filter()
            if date_from and date_to:
                clauses.append("date_de_reception BETWEEN ? AND ?"); params.extend([date_from, date_to])
            elif date_from:
                clauses.append("date_de_reception >= ?"); params.append(date_from)
            elif date_to:
                clauses.append("date_de_reception <= ?"); params.append(date_to)

        elif self._table_name == "join_employeur_traitement":
            if (v := self.localite_combo.currentData()) is not None:
                clauses.append("ie.localites = ?"); params.append(v)
            if (v := self.exercice_combo.currentData()) is not None:
                clauses.append("(td.exercice = ? OR ie.exercice = ?)"); params.extend([v, v])
            if (v := self.secteur_combo.currentData()) is not None:
                # "NON TRAITÉ" est stocké NULL en base — on filtre sur IS NULL
                if str(v).upper() in ("NON TRAITÉ", "NON TRAITE"):
                    clauses.append("(td.statut IS NULL OR td.statut = 'NON TRAITÉ')")
                else:
                    clauses.append("td.statut = ?"); params.append(v)
            date_from, date_to = self._get_date_filter()
            if date_from and date_to:
                clauses.append("td.date_de_reception BETWEEN ? AND ?"); params.extend([date_from, date_to])
            elif date_from:
                clauses.append("td.date_de_reception >= ?"); params.append(date_from)
            elif date_to:
                clauses.append("td.date_de_reception <= ?"); params.append(date_to)
            # Filtre suspension (commun à la vue jointe)
            susp_val = self.suspension_combo.currentData()
            if susp_val is not None:
                clauses.append("COALESCE(td.is_suspended, 0) = ?"); params.append(int(susp_val))

        # Filtre suspension pour les autres tables
        if self._table_name == "traitement_disa":
            susp_val = self.suspension_combo.currentData()
            if susp_val is not None:
                clauses.append("COALESCE(is_suspended, 0) = ?"); params.append(int(susp_val))

        where = (" WHERE " + " AND ".join(clauses)) if clauses else ""
        return where, params

    def _build_select_sql(self, base_from: str, where: str, extra: str = "") -> str:
        """Construit le SELECT en cohérence avec _load_structure."""
        if self._table_name == "join_employeur_traitement":
            try:
                conn = get_connection()
                with conn:
                    cur = conn.cursor()
                    cur.execute("PRAGMA table_info(identification_employeurs)")
                    emp_cols = [r[1] for r in cur.fetchall()]
                    cur.execute("PRAGMA table_info(traitement_disa)")
                    td_cols = [r[1] for r in cur.fetchall()]
            except Exception:
                emp_cols, td_cols = [], []

            duplicates = set(emp_cols) & set(td_cols)
            parts: list[str] = ["ie.id AS id_employeur"]
            for name in emp_cols:
                if name != "id":
                    parts.append(f"ie.{name} AS {name}")
            parts.append("td.id AS id_traitement")
            for name in td_cols:
                if name == "id":
                    continue
                alias = "statut" if name == "statut" else (f"{name}_disa" if name in duplicates else name)
                if name == "statut":
                    expr = "COALESCE(td.statut, 'NON TRAITÉ')"
                elif name == "is_suspended":
                    expr = "COALESCE(td.is_suspended, 0)"
                else:
                    expr = f"td.{name}"
                parts.append(f"{expr} AS {alias}")
            return "SELECT " + ", ".join(parts) + base_from + where + extra
        else:
            return (
                "SELECT " + ", ".join(self._columns) + base_from + where + extra
            )

    # ------------------------------------------------------------------
    # Rafraîchissement du tableau
    # ------------------------------------------------------------------

    def _refresh_table(self) -> None:
        if not self._columns:
            return
        # Ne pas recharger si le widget n'est pas affiché (évite de bloquer le thread
        # principal lors du polling data_changed toutes les 4 s)
        if not self.isVisible():
            self._needs_refresh = True
            return
        self._needs_refresh = False

        where, params = self._build_filters_sql()

        base_from = (
            " FROM identification_employeurs ie "
            "LEFT JOIN traitement_disa td ON td.employeur_id = ie.id"
            if self._table_name == "join_employeur_traitement"
            else f" FROM {self._table_name}"
        )

        count_sql = "SELECT COUNT(*)" + base_from + where

        try:
            conn = get_connection()
            with conn:
                cur = conn.cursor()
                cur.execute(count_sql, params)
                (self._total_rows,) = cur.fetchone() or (0,)

                # Compteurs par statut
                if self._table_name in ("traitement_disa", "join_employeur_traitement"):
                    if self._table_name == "traitement_disa":
                        group_sql = (
                            f"SELECT statut, COUNT(*){base_from}{where} GROUP BY statut"
                        )
                    else:
                        group_sql = (
                            "SELECT COALESCE(td.statut, 'NON TRAITÉ') AS statut, COUNT(*)"
                            + base_from + where
                            + " GROUP BY COALESCE(td.statut, 'NON TRAITÉ')"
                        )
                    cur.execute(group_sql, params)
                    non_traite = valide = rejet = 0
                    for statut, nb in cur.fetchall():
                        txt = str(statut).upper()
                        nb_i = int(nb or 0)
                        if "NON" in txt and "TRAIT" in txt:
                            non_traite += nb_i
                        elif "REJET" in txt:
                            rejet += nb_i
                        else:
                            valide += nb_i
                    self.status_non_traite_lbl.setText(f"Non traités : {non_traite}")
                    self.status_valide_lbl.setText(f"Validés : {valide}")
                    self.status_rejet_lbl.setText(f"Avec rejets : {rejet}")
                else:
                    self.status_non_traite_lbl.setText("Non traités : 0")
                    self.status_valide_lbl.setText("Validés : 0")
                    self.status_rejet_lbl.setText("Avec rejets : 0")

                total_pages = max(1, (self._total_rows + self._page_size - 1) // self._page_size)
                if self._current_page > total_pages:
                    self._current_page = total_pages

                offset = (self._current_page - 1) * self._page_size
                order = " ORDER BY ie.id DESC" if self._table_name == "join_employeur_traitement" else " ORDER BY id DESC"
                sql = self._build_select_sql(base_from, where, f"{order} LIMIT ? OFFSET ?")
                cur.execute(sql, params + [self._page_size, offset])
                rows = cur.fetchall()
        except Exception as exc:
            QMessageBox.critical(
                self, "Erreur base de données",
                f"Impossible de charger les enregistrements :\n{exc}",
            )
            return

        self.table.setRowCount(len(rows))
        statut_col_index = (
            self._columns.index("statut")
            if self._table_name in ("traitement_disa", "join_employeur_traitement") and "statut" in self._columns
            else -1
        )
        suspended_col_index = (
            self._columns.index("is_suspended")
            if "is_suspended" in self._columns
            else -1
        )

        for row_index, row in enumerate(rows):
            bg_color = None
            fg_color = QColor("white")
            etat_text = ""
            etat_icon: QIcon | None = None
            is_suspended_val = False

            if statut_col_index != -1:
                try:
                    row_statut = str(row[statut_col_index] or "")
                except Exception:
                    row_statut = ""
                upper = row_statut.upper()
                if "NON" in upper and "TRAIT" in upper:
                    bg_color = QColor("#fee2e2"); fg_color = QColor("#991b1b")
                    etat_text = "⊘  Non traité"
                    etat_icon = QIcon(":/icon/icon/close-window-64.ico")
                elif "REJET" in upper:
                    bg_color = QColor("#fef3c7"); fg_color = QColor("#92400e")
                    etat_text = "⚠  Avec rejets"
                    etat_icon = QIcon(":/icon/icon/activity-feed-32.ico")
                elif "VALID" in upper or "TRAIT" in upper:
                    bg_color = QColor("#dcfce7"); fg_color = QColor("#14532d")
                    etat_text = "✔  Validé"
                    etat_icon = QIcon(":/icon/icon/dashboard-5-32.ico")
                else:
                    etat_text = row_statut

            if suspended_col_index != -1:
                try:
                    is_suspended_val = bool(int(row[suspended_col_index] or 0))
                except (TypeError, ValueError):
                    is_suspended_val = False

            # Colonne 0 : case à cocher
            cb_item = QTableWidgetItem()
            cb_item.setFlags(
                Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled
            )
            cb_item.setCheckState(Qt.CheckState.Unchecked)
            cb_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_index, 0, cb_item)

            for col_index, value in enumerate(row):
                # Afficher "Oui"/"Non" pour la colonne is_suspended
                if col_index == suspended_col_index:
                    text = "Oui" if bool(int(value or 0)) else "Non"
                else:
                    text = "" if value is None else str(value)
                item = QTableWidgetItem(text)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if self._id_index != -1 and col_index == self._id_index:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if col_index == suspended_col_index:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    if is_suspended_val:
                        item.setForeground(QColor("#b45309"))
                        item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
                    else:
                        item.setForeground(QColor("#15803d"))
                elif bg_color is not None and statut_col_index != -1:
                    item.setBackground(bg_color)
                    item.setForeground(fg_color)
                self.table.setItem(row_index, col_index + 1, item)  # +1 : offset checkbox

            if self._table_name in ("traitement_disa", "join_employeur_traitement"):
                etat_item = QTableWidgetItem(etat_text)
                etat_item.setFlags(etat_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if etat_icon is not None:
                    etat_item.setIcon(etat_icon)
                if bg_color is not None:
                    etat_item.setBackground(bg_color)
                    etat_item.setForeground(fg_color)
                self.table.setItem(row_index, len(self._columns) + 1, etat_item)

        # Mettre à jour le bouton Suspendre selon la ligne sélectionnée
        self._update_suspend_button()
        self._update_delete_btn_state()
        self.table.resizeColumnsToContents()
        self.table.setColumnWidth(0, 36)  # Colonne checkbox — largeur fixe

        total_pages = max(1, (self._total_rows + self._page_size - 1) // self._page_size)
        self.page_label.setText(
            f"Page {self._current_page} / {total_pages}  —  {self._total_rows} ligne(s)"
        )
        self.prev_page_btn.setEnabled(self._current_page > 1)
        self.next_page_btn.setEnabled(self._current_page < total_pages)

    # ------------------------------------------------------------------
    # Export Excel
    # ------------------------------------------------------------------

    def _on_export_excel_clicked(self) -> None:
        """Exporte toutes les lignes correspondant aux filtres actifs vers Excel."""

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(
                self, "Module manquant",
                "Le module 'openpyxl' est requis pour l'export Excel.\n"
                "Installez-le avec : pip install openpyxl",
            )
            return

        if not self._columns:
            QMessageBox.warning(self, "Export", "Aucune colonne à exporter.")
            return

        # Choix du fichier de destination
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Enregistrer l'export Excel",
            "export_base_disa.xlsx",
            "Fichiers Excel (*.xlsx)",
        )
        if not file_path:
            return

        where, params = self._build_filters_sql()
        base_from = (
            " FROM identification_employeurs ie "
            "LEFT JOIN traitement_disa td ON td.employeur_id = ie.id"
            if self._table_name == "join_employeur_traitement"
            else f" FROM {self._table_name}"
        )

        order = (
            " ORDER BY ie.id ASC"
            if self._table_name == "join_employeur_traitement"
            else " ORDER BY id ASC"
        )
        sql = self._build_select_sql(base_from, where, order)

        try:
            conn = get_connection()
            with conn:
                cur = conn.cursor()
                cur.execute(sql, params)
                all_rows = cur.fetchall()
        except Exception as exc:
            QMessageBox.critical(
                self, "Erreur base de données",
                f"Impossible de récupérer les données :\n{exc}",
            )
            return

        # Construire les en-têtes
        if self._table_name in ("traitement_disa", "join_employeur_traitement"):
            headers = list(self._columns) + ["État"]
        else:
            headers = list(self._columns)

        # ── Construire le workbook ─────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export DISA"

        # Styles d'en-tête
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border_side = Side(style="thin", color="D1D5DB")
        thin_border = Border(
            left=thin_border_side, right=thin_border_side,
            top=thin_border_side, bottom=thin_border_side,
        )

        # Ligne d'en-tête
        for col_num, col_name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name.replace("_", " ").upper())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[1].height = 22

        # Fills pour les statuts
        fill_red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        fill_green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        fill_orange = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        font_red = Font(color="991B1B")
        font_green = Font(color="14532D")
        font_orange = Font(color="92400E")

        statut_col_index = (
            self._columns.index("statut")
            if self._table_name in ("traitement_disa", "join_employeur_traitement") and "statut" in self._columns
            else -1
        )

        # Données
        for row_num, row in enumerate(all_rows, 2):
            # Détecter le statut pour colorier la ligne
            row_fill = None
            row_font_col = None
            if statut_col_index != -1:
                try:
                    statut_val = str(row[statut_col_index] or "").upper()
                except Exception:
                    statut_val = ""
                if "NON" in statut_val and "TRAIT" in statut_val:
                    row_fill, row_font_col = fill_red, font_red
                    etat_label = "Non traité"
                elif "REJET" in statut_val:
                    row_fill, row_font_col = fill_orange, font_orange
                    etat_label = "Avec rejets"
                elif "VALID" in statut_val or "TRAIT" in statut_val:
                    row_fill, row_font_col = fill_green, font_green
                    etat_label = "Validé"
                else:
                    etat_label = statut_val
            else:
                etat_label = ""

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=(None if value is None else str(value)))
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
                    cell.font = row_font_col
                cell.alignment = Alignment(vertical="center")

            # Colonne "État" si nécessaire
            if self._table_name in ("traitement_disa", "join_employeur_traitement"):
                etat_cell = ws.cell(row=row_num, column=len(self._columns) + 1, value=etat_label)
                etat_cell.border = thin_border
                if row_fill:
                    etat_cell.fill = row_fill
                    etat_cell.font = row_font_col

            ws.row_dimensions[row_num].height = 16

        # Ajuster la largeur des colonnes
        for col_num, col_name in enumerate(headers, 1):
            max_len = max(len(col_name), 10)
            for row_num in range(2, min(len(all_rows) + 2, 52)):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_num)
            ].width = min(max_len + 3, 40)

        # Figer la ligne d'en-tête
        ws.freeze_panes = "A2"

        # Filtres automatiques Excel sur la ligne d'en-tête
        ws.auto_filter.ref = ws.dimensions

        try:
            wb.save(file_path)
        except Exception as exc:
            QMessageBox.critical(
                self, "Erreur d'export",
                f"Impossible d'enregistrer le fichier Excel :\n{exc}",
            )
            return

        QMessageBox.information(
            self, "Export réussi",
            f"Export terminé avec succès !\n\n"
            f"• Lignes exportées : {len(all_rows)}\n"
            f"• Fichier : {file_path}",
        )

    # ------------------------------------------------------------------
    # Pagination
    # ------------------------------------------------------------------

    def _on_prev_page(self) -> None:
        if self._current_page > 1:
            self._current_page -= 1
            self._refresh_table()

    def _on_next_page(self) -> None:
        total_pages = max(1, (self._total_rows + self._page_size - 1) // self._page_size)
        if self._current_page < total_pages:
            self._current_page += 1
            self._refresh_table()

    def _on_page_size_changed(self) -> None:
        try:
            new_size = int(self.page_size_combo.currentText())
        except ValueError:
            return
        if new_size > 0:
            self._page_size = new_size
            self._current_page = 1
            self._refresh_table()

    # ------------------------------------------------------------------
    # Actions CRUD
    # ------------------------------------------------------------------

    def _get_selected_ids(self) -> list[int]:
        """Retourne les IDs de toutes les lignes dont la case est cochée."""
        if self._table_name == "join_employeur_traitement":
            if "id_employeur" not in self._columns:
                return []
            col_idx = self._columns.index("id_employeur") + 1  # +1 checkbox
        elif self._id_index != -1:
            col_idx = self._id_index + 1  # +1 checkbox
        else:
            return []
        ids = []
        for row in range(self.table.rowCount()):
            cb = self.table.item(row, 0)
            if cb and cb.checkState() == Qt.CheckState.Checked:
                item = self.table.item(row, col_idx)
                if item and item.text().strip():
                    try:
                        ids.append(int(item.text().strip()))
                    except ValueError:
                        pass
        return ids

    def _update_delete_btn_state(self) -> None:
        """Active le bouton Supprimer seulement si au moins une case est cochée."""
        has_checked = any(
            self.table.item(r, 0) is not None
            and self.table.item(r, 0).checkState() == Qt.CheckState.Checked
            for r in range(self.table.rowCount())
        )
        self.delete_btn.setEnabled(has_checked)

    def _on_checkbox_changed(self, item: QTableWidgetItem) -> None:
        """Réagit au changement d'état d'une case à cocher."""
        if item.column() == 0:
            self._update_delete_btn_state()

    def _on_header_section_clicked(self, col: int) -> None:
        """Clic sur l'en-tête de la colonne 0 → cocher / décocher tout."""
        if col != 0:
            return
        any_unchecked = any(
            self.table.item(r, 0) is not None
            and self.table.item(r, 0).checkState() == Qt.CheckState.Unchecked
            for r in range(self.table.rowCount())
        )
        new_state = Qt.CheckState.Checked if any_unchecked else Qt.CheckState.Unchecked
        self.table.blockSignals(True)
        for r in range(self.table.rowCount()):
            cb = self.table.item(r, 0)
            if cb:
                cb.setCheckState(new_state)
        self.table.blockSignals(False)
        self._update_delete_btn_state()

    def _get_selected_employeur_id(self) -> Optional[int]:
        row = self.table.currentRow()
        if row < 0 or "id_employeur" not in self._columns:
            return None
        item = self.table.item(row, self._columns.index("id_employeur") + 1)  # +1 checkbox
        if item is None:
            return None
        try:
            return int(item.text())
        except ValueError:
            return None

    def _get_selected_id(self) -> Optional[int]:
        row = self.table.currentRow()
        if row < 0 or self._id_index == -1:
            return None
        item = self.table.item(row, self._id_index + 1)  # +1 checkbox
        if item is None:
            return None
        try:
            return int(item.text())
        except ValueError:
            return None

    def _collect_row_data(self, row: int) -> Dict[str, Optional[str]]:
        data: Dict[str, Optional[str]] = {}
        for col_index, col_name in enumerate(self._columns):
            item = self.table.item(row, col_index + 1)  # +1 checkbox
            text = "" if item is None else item.text().strip()
            data[col_name] = text if text else None
        return data

    def _on_add_clicked(self) -> None:
        if not self._columns:
            return

        if self._table_name == "join_employeur_traitement":
            try:
                conn = get_connection()
                with conn:
                    cur = conn.cursor()
                    cur.execute("PRAGMA table_info(identification_employeurs)")
                    emp_cols = [r[1] for r in cur.fetchall() if r[1] != "id"]
            except Exception as exc:
                QMessageBox.critical(self, "Erreur base de données", str(exc)); return

            dialog = EmployeurFormDialog(self, emp_cols)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                values = dialog.get_values()
                cols = list(values.keys())
                placeholders = ", ".join(["?"] * len(cols))
                sql = (
                    f"INSERT INTO identification_employeurs ({', '.join(cols)}) "
                    f"VALUES ({placeholders})"
                )
                try:
                    conn = get_connection()
                    with conn:
                        conn.cursor().execute(sql, [values[c] for c in cols])
                except Exception as exc:
                    QMessageBox.critical(self, "Erreur base de données", str(exc)); return
                self._load_filters(); self._refresh_table()
                get_data_bus().data_changed.emit()
            return

        dialog = EmployeurFormDialog(self, self._columns)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            values = dialog.get_values()
            cols = list(values.keys())
            placeholders = ", ".join(["?"] * len(cols))
            sql = f"INSERT INTO {self._table_name} ({', '.join(cols)}) VALUES ({placeholders})"
            try:
                conn = get_connection()
                with conn:
                    conn.cursor().execute(sql, [values[c] for c in cols])
            except Exception as exc:
                QMessageBox.critical(self, "Erreur base de données", str(exc)); return
            self._load_filters(); self._refresh_table()

    def _on_edit_clicked(self) -> None:
        if not self._columns:
            return

        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Modification", "Sélectionnez d'abord une ligne à modifier.")
            return

        if self._table_name == "join_employeur_traitement":
            emp_id = self._get_selected_employeur_id()
            if emp_id is None:
                QMessageBox.warning(self, "Modification", "Impossible de récupérer l'identifiant de l'employeur.")
                return
            try:
                conn = get_connection()
                with conn:
                    cur = conn.cursor()
                    cur.execute("PRAGMA table_info(identification_employeurs)")
                    emp_cols = [r[1] for r in cur.fetchall() if r[1] != "id"]
                    cur.execute(
                        "SELECT " + ", ".join(emp_cols) + " FROM identification_employeurs WHERE id = ?",
                        (emp_id,),
                    )
                    db_row = cur.fetchone()
            except Exception as exc:
                QMessageBox.critical(self, "Erreur base de données", str(exc)); return

            if db_row is None:
                QMessageBox.warning(self, "Modification", "L'employeur sélectionné n'existe plus."); return

            current_data = {col: (None if v is None else str(v)) for col, v in zip(emp_cols, db_row)}
            dialog = EmployeurFormDialog(self, emp_cols, current_data)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                values = dialog.get_values()
                set_clauses = [f"{col} = ?" for col in values]
                params = list(values.values()) + [emp_id]
                sql = "UPDATE identification_employeurs SET " + ", ".join(set_clauses) + " WHERE id = ?"
                try:
                    conn = get_connection()
                    with conn:
                        conn.cursor().execute(sql, params)
                except Exception as exc:
                    QMessageBox.critical(self, "Erreur base de données", str(exc)); return
                self._load_filters(); self._refresh_table()
                get_data_bus().data_changed.emit()
            return

        emp_id = self._get_selected_id()
        if emp_id is None:
            QMessageBox.warning(self, "Modification", "Impossible de récupérer l'identifiant."); return
        current_data = self._collect_row_data(row)
        dialog = EmployeurFormDialog(self, self._columns, current_data)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            values = dialog.get_values()
            set_clauses = [f"{col} = ?" for col in values]
            params = list(values.values()) + [emp_id]
            sql = f"UPDATE {self._table_name} SET " + ", ".join(set_clauses) + " WHERE id = ?"
            try:
                conn = get_connection()
                with conn:
                    conn.cursor().execute(sql, params)
            except Exception as exc:
                QMessageBox.critical(self, "Erreur base de données", str(exc)); return
            self._load_filters(); self._refresh_table()

    # ------------------------------------------------------------------
    # Suspension d'entreprise
    # ------------------------------------------------------------------

    def _get_current_suspension_state(self) -> Optional[bool]:
        """Retourne True si l'entreprise sélectionnée est suspendue, False sinon,
        None si aucune ligne sélectionnée ou pas de fiche traitement DISA associée."""
        if self._table_name not in ("join_employeur_traitement", "traitement_disa"):
            return None
        if "is_suspended" not in self._columns:
            return None
        row = self.table.currentRow()
        if row < 0:
            return None

        # Vérifier qu'un traitement DISA existe pour cette ligne
        # (LEFT JOIN peut retourner NULL pour td.id si l'employeur n'a pas de fiche)
        if "id_traitement" in self._columns:
            id_col = self._columns.index("id_traitement") + 1  # +1 checkbox
            id_item = self.table.item(row, id_col)
            if id_item is None or not id_item.text().strip():
                return None  # Pas de fiche traitement → suspension impossible

        col_idx = self._columns.index("is_suspended") + 1  # +1 checkbox
        item = self.table.item(row, col_idx)
        if item is None:
            return None
        # La cellule affiche "Oui"/"Non"
        return item.text().strip().lower() == "oui"

    def _update_suspend_button(self) -> None:
        """Met à jour le libellé et le style du bouton selon l'état de suspension."""
        state = self._get_current_suspension_state()
        if state is None:
            self.suspend_btn.setText("⛔  Suspendre")
            self.suspend_btn.setStyleSheet(_STYLE_BTN_SUSPEND)
        elif state:
            # Déjà suspendue → proposer de réactiver
            self.suspend_btn.setText("✔  Réactiver")
            self.suspend_btn.setStyleSheet(_STYLE_BTN_REACTIVATE)
        else:
            self.suspend_btn.setText("⛔  Suspendre")
            self.suspend_btn.setStyleSheet(_STYLE_BTN_SUSPEND)

    def _on_toggle_suspension_clicked(self) -> None:
        """Bascule l'état is_suspended de l'entreprise sélectionnée."""
        if self._table_name not in ("join_employeur_traitement", "traitement_disa"):
            QMessageBox.information(
                self, "Suspension",
                "La suspension n'est disponible que sur la vue Employeurs + DISA."
            )
            return

        state = self._get_current_suspension_state()
        if state is None:
            QMessageBox.information(
                self, "Suspension",
                "Cette entreprise n'a pas encore de fiche de traitement DISA.\n"
                "Veuillez d'abord créer une fiche de traitement pour pouvoir la suspendre."
            )
            return

        new_state = 0 if state else 1
        label = "suspendre" if new_state else "réactiver"
        label_past = "suspendue" if new_state else "réactivée"

        # Récupérer l'id_traitement (colonne "id_traitement" dans la vue jointe)
        traitement_id: Optional[int] = None
        if "id_traitement" in self._columns:
            col_idx = self._columns.index("id_traitement") + 1  # +1 checkbox
            item = self.table.item(self.table.currentRow(), col_idx)
            if item and item.text().strip():
                try:
                    traitement_id = int(item.text().strip())
                except ValueError:
                    pass

        if traitement_id is None:
            QMessageBox.warning(
                self, "Suspension",
                "Impossible de trouver l'identifiant du traitement DISA pour cette ligne."
            )
            return

        reply = QMessageBox.question(
            self, f"Confirmer — {label.capitalize()}",
            f"Voulez-vous vraiment {label} cette entreprise ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            conn = get_connection()
            with conn:
                conn.execute(
                    "UPDATE traitement_disa SET is_suspended = ?, updated_at = datetime('now') WHERE id = ?",
                    (new_state, traitement_id),
                )
        except Exception as exc:
            QMessageBox.critical(
                self, "Erreur base de données",
                f"Impossible de {label} l'entreprise :\n{exc}"
            )
            return

        QMessageBox.information(
            self, "Succès",
            f"L'entreprise a bien été {label_past}."
        )
        self._refresh_table()
        get_data_bus().data_changed.emit()

    def _on_delete_clicked(self) -> None:
        if not self._columns:
            return

        ids = self._get_selected_ids()
        if not ids:
            QMessageBox.information(self, "Suppression", "Sélectionnez d'abord une ou plusieurs lignes à supprimer.")
            return

        count = len(ids)
        if count == 1:
            msg = "Voulez-vous vraiment supprimer cet enregistrement ?\n(Cette action est irréversible)"
        else:
            msg = f"Voulez-vous vraiment supprimer ces {count} enregistrements ?\n(Cette action est irréversible)"

        reply = QMessageBox.question(
            self, "Confirmer la suppression",
            msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        table = "identification_employeurs" if self._table_name == "join_employeur_traitement" else self._table_name
        placeholders = ",".join("?" * count)
        try:
            conn = get_connection()
            with conn:
                conn.cursor().execute(
                    f"DELETE FROM {table} WHERE id IN ({placeholders})", ids
                )
        except Exception as exc:
            QMessageBox.critical(self, "Erreur base de données", str(exc))
            return

        self._load_filters()
        self._refresh_table()
        get_data_bus().data_changed.emit()
