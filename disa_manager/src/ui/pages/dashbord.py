from __future__ import annotations

import logging

from PySide6.QtCore import Qt, QMargins, QPropertyAnimation, QEasingCurve, QTimer, QDate

logger = logging.getLogger(__name__)
from PySide6.QtGui import QPainter, QColor, QFont, QCursor
from PySide6.QtCharts import (
    QBarCategoryAxis,
    QBarSeries,
    QBarSet,
    QChart,
    QChartView,
    QHorizontalBarSeries,
    QPieSeries,
    QValueAxis,
)
from PySide6.QtWidgets import (
    QGridLayout, QLabel, QFrame, QVBoxLayout, QHBoxLayout,
    QDialog, QGraphicsOpacityEffect, QPushButton, QDateEdit,
)

from db.connection import get_connection
from core.events import get_data_bus
from ui.dashboard_theme import (
    _BG_DEEP, _BG_CARD, _BG_HDR, _BORDER,
    _TXT1, _TXT2, _TXT3, _GRID,
    _C_NAVY, _C_GREEN, _C_AMBER, _C_BLUE,
    _USER_COLORS, DATE_EDIT_QSS as _DATE_EDIT_QSS, BTN_QSS as _BTN_QSS,
    truncate as _truncate,
)


class ChartWidget:
    def __init__(self, widget):
        super().__init__()
        self.widget = widget
        self._responsive_label_specs: list[tuple[QLabel, int]] = []
        self._detail_dialog: QDialog | None = None
        self._detail_label: QLabel | None = None
        self._animations: list[QPropertyAnimation] = []
        self._last_data: dict = {}  # Cache des dernières données pour l'export

        # Filtres actifs (plage de dates — défaut : année en cours)
        _y = QDate.currentDate().year()
        self._filter_date_from: str = f"{_y}-01-01"
        self._filter_date_to:   str = f"{_y}-12-31"

        get_data_bus().data_changed.connect(self.refresh)

    def refresh(self) -> None:
        self._animations.clear()
        self.add_chart()

    # ── Filtres ───────────────────────────────────────────────────────────────

    def _date_conditions(self) -> tuple[str, str, list]:
        """Renvoie (join_cond, and_cond, params) pour la plage de dates active.

        Filtre sur COALESCE(updated_at, created_at) :
        - updated_at est mis à jour à chaque modification (via le formulaire)
        - created_at est la date d'insertion initiale (import Excel)
        → les enregistrements modifiés sont toujours inclus dans le filtre
        join_cond : dans la clause ON du LEFT JOIN (préserve le LEFT JOIN)
        and_cond  : dans un WHERE existant (avec AND)
        """
        date_col = "COALESCE(td.updated_at, td.created_at)"
        join_cond = f"AND date({date_col}) BETWEEN ? AND ?"
        and_cond  = f"AND date({date_col}) BETWEEN ? AND ?"
        return join_cond, and_cond, [self._filter_date_from, self._filter_date_to]

    def _make_filter_bar(self) -> QFrame:
        """Barre de filtres sombre (plage de dates : Du … Au …)."""
        bar = QFrame()
        bar.setStyleSheet(
            f"""
            QFrame {{
                background-color: {_BG_HDR};
                border-radius: 10px;
                border: 1px solid {_BORDER};
            }}
            QLabel {{
                color: {_TXT2};
                font-size: 11px;
                font-weight: 600;
                background: transparent;
                border: none;
            }}
            """
        )
        bar.setMinimumHeight(46)

        hbox = QHBoxLayout(bar)
        hbox.setContentsMargins(16, 8, 16, 8)
        hbox.setSpacing(10)

        icon = QLabel("🗓")
        icon.setStyleSheet("font-size: 14px; border: none; background: transparent;")
        hbox.addWidget(icon)

        def _make_date_edit(stored_value: str) -> QDateEdit:
            de = QDateEdit()
            de.setCalendarPopup(True)
            de.setDisplayFormat("dd/MM/yyyy")
            de.setStyleSheet(_DATE_EDIT_QSS)
            de.setDate(QDate.fromString(stored_value, "yyyy-MM-dd"))
            return de

        hbox.addWidget(QLabel("Du :"))
        date_from = _make_date_edit(self._filter_date_from)
        date_from.dateChanged.connect(self._on_date_from_changed)
        hbox.addWidget(date_from)

        hbox.addWidget(QLabel("  Au :"))
        date_to = _make_date_edit(self._filter_date_to)
        date_to.dateChanged.connect(self._on_date_to_changed)
        hbox.addWidget(date_to)

        hbox.addStretch(1)

        reset_btn = QPushButton("↺  Réinitialiser")
        reset_btn.setStyleSheet(_BTN_QSS)
        reset_btn.clicked.connect(self._on_filter_reset)
        hbox.addWidget(reset_btn)

        export_btn = QPushButton("⬇  Exporter Excel")
        export_btn.setStyleSheet(
            "QPushButton {"
            "  background-color: #15803d; color: white; border-radius: 5px;"
            "  padding: 5px 14px; font-size: 11px; font-weight: 600; border: none;"
            "}"
            "QPushButton:hover  { background-color: #16a34a; }"
            "QPushButton:pressed { background-color: #0f5c2c; }"
        )
        export_btn.setToolTip("Exporter le tableau de bord dans un fichier Excel")
        export_btn.clicked.connect(self._export_to_excel)
        hbox.addWidget(export_btn)

        return bar

    def _on_date_from_changed(self, date: QDate) -> None:
        self._filter_date_from = date.toString("yyyy-MM-dd")
        self.refresh()

    def _on_date_to_changed(self, date: QDate) -> None:
        self._filter_date_to = date.toString("yyyy-MM-dd")
        self.refresh()

    def _on_filter_reset(self) -> None:
        _y = QDate.currentDate().year()
        self._filter_date_from = f"{_y}-01-01"
        self._filter_date_to   = f"{_y}-12-31"
        self.refresh()

    # ── Helpers UI ────────────────────────────────────────────────────────────

    def _fade_in(self, widget, delay_ms: int = 0) -> None:
        effect = QGraphicsOpacityEffect(widget)
        effect.setOpacity(0.0)
        widget.setGraphicsEffect(effect)

        def _start() -> None:
            anim = QPropertyAnimation(effect, b"opacity")
            anim.setStartValue(0.0)
            anim.setEndValue(1.0)
            anim.setDuration(550)
            anim.setEasingCurve(QEasingCurve.Type.OutCubic)
            anim.start(QPropertyAnimation.DeletionPolicy.DeleteWhenStopped)
            self._animations.append(anim)

        if delay_ms > 0:
            QTimer.singleShot(delay_ms, _start)
        else:
            _start()

    def _make_kpi_card(
        self, title: str, value: str, subtitle: str, accent: str, delay_ms: int = 0,
    ) -> QFrame:
        frame = QFrame()
        frame.setStyleSheet(
            f"""
            QFrame {{
                background-color: {_BG_CARD};
                border-radius: 12px;
                border: 1px solid {_BORDER};
                border-top: 3px solid {accent};
            }}
            """
        )
        frame.setMinimumHeight(110)
        vbox = QVBoxLayout(frame)
        vbox.setContentsMargins(14, 10, 14, 10)
        vbox.setSpacing(4)

        lbl_title = QLabel(title.upper())
        lbl_title.setStyleSheet(
            f"color: {_TXT3}; font-weight: 600; letter-spacing: 0.7px;"
            " border: none; background: transparent;"
        )
        lbl_value = QLabel(value)
        lbl_value.setStyleSheet(
            f"color: {accent}; font-size: 20px; font-weight: 800;"
            " border: none; background: transparent;"
        )
        lbl_sub = QLabel(subtitle)
        lbl_sub.setStyleSheet(
            f"color: {_TXT3}; border: none; background: transparent;"
        )
        vbox.addWidget(lbl_title)
        vbox.addWidget(lbl_value)
        vbox.addWidget(lbl_sub)
        vbox.addStretch(1)

        self._responsive_label_specs.extend([
            (lbl_title, 9), (lbl_value, 18), (lbl_sub, 9),
        ])
        self._fade_in(frame, delay_ms)
        return frame

    def _make_chart_card(
        self, title: str, subtitle: str, delay_ms: int = 0,
    ) -> tuple["QFrame", "QVBoxLayout"]:
        card = QFrame()
        card.setStyleSheet(
            f"""
            QFrame {{
                background-color: {_BG_CARD};
                border-radius: 12px;
                border: 1px solid {_BORDER};
            }}
            """
        )
        outer = QVBoxLayout(card)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        hdr = QFrame()
        hdr.setStyleSheet(
            f"""
            QFrame {{
                background-color: {_BG_HDR};
                border: none;
                border-radius: 12px;
                border-bottom-left-radius: 0;
                border-bottom-right-radius: 0;
                border-bottom: 1px solid {_BORDER};
            }}
            """
        )
        hdr_lay = QVBoxLayout(hdr)
        hdr_lay.setContentsMargins(18, 11, 18, 9)
        hdr_lay.setSpacing(2)

        lbl_t = QLabel(title)
        lbl_t.setStyleSheet(
            f"color: {_TXT1}; font-weight: 700; font-size: 12px;"
            " border: none; background: transparent;"
        )
        lbl_s = QLabel(subtitle)
        lbl_s.setStyleSheet(
            f"color: {_TXT2}; font-size: 10px; border: none; background: transparent;"
        )
        hdr_lay.addWidget(lbl_t)
        hdr_lay.addWidget(lbl_s)
        outer.addWidget(hdr)

        self._fade_in(card, delay_ms)
        return card, outer

    def _base_chart(self, bg: str = _BG_CARD) -> QChart:
        chart = QChart()
        chart.setTheme(QChart.ChartTheme.ChartThemeDark)
        chart.setTitle("")
        chart.setBackgroundBrush(QColor(bg))
        chart.setBackgroundRoundness(0)
        chart.setMargins(QMargins(10, 6, 10, 6))
        chart.setAnimationOptions(QChart.AnimationOption.AllAnimations)
        chart.setAnimationDuration(700)
        tf = chart.titleFont()
        tf.setPointSize(1)
        chart.setTitleFont(tf)
        return chart

    @staticmethod
    def _styled_chart_view(chart: QChart, min_h: int = 220) -> QChartView:
        view = QChartView(chart)
        view.setRenderHint(QPainter.Antialiasing)
        view.setStyleSheet("background: transparent; border: none;")
        view.setMinimumHeight(min_h)
        return view

    def _style_legend(self, chart: QChart) -> None:
        chart.legend().setVisible(True)
        chart.legend().setAlignment(Qt.AlignBottom)
        lf = chart.legend().font()
        lf.setPointSize(9)
        chart.legend().setFont(lf)
        chart.legend().setColor(QColor(_TXT1))
        chart.legend().setLabelColor(QColor(_TXT2))

    def _style_axis(self, axis) -> None:
        lf = axis.labelsFont()
        lf.setPointSize(9)
        axis.setLabelsFont(lf)
        axis.setLabelsColor(QColor(_TXT2))
        if isinstance(axis, QValueAxis):
            axis.setGridLineColor(QColor(_GRID))

    # ── Construction principale ───────────────────────────────────────────────

    def add_chart(self):
        layout: QGridLayout = self.widget
        self._responsive_label_specs = []
        self._animations.clear()

        while layout.count():
            item = layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()

        parent = layout.parentWidget()
        if parent is not None:
            parent.setStyleSheet(f"background-color: {_BG_DEEP};")

        # ── Données ───────────────────────────────────────────────────
        join_cond, and_cond, dparams = self._date_conditions()

        total_employeurs = 0
        lignes_traitees = 0
        lignes_non_traitees = 0
        lignes_suspendues = 0
        localites: list[str] = []
        disa_restantes: list[int] = []
        disa_traitees_loc: list[int] = []
        secteurs: list[str] = []
        secteurs_nb: list[int] = []
        traite_par_users: list[str] = []
        traite_par_nb: list[int] = []

        try:
            conn = get_connection()
            with conn:
                cur = conn.cursor()

                # base_from sans filtre (total global)
                base_from = (
                    " FROM identification_employeurs ie "
                    "LEFT JOIN traitement_disa td ON td.employeur_id = ie.id"
                )
                # base_from avec filtre dans le JOIN (préserve le LEFT JOIN)
                base_from_f = (
                    " FROM identification_employeurs ie "
                    f"LEFT JOIN traitement_disa td ON td.employeur_id = ie.id {join_cond}"
                )

                # Total employeurs — non filtré (comptage global)
                cur.execute("SELECT COUNT(DISTINCT ie.id)" + base_from)
                total_employeurs = cur.fetchone()[0] or 0

                # DISA suspendues (priorité sur le statut)
                cur.execute(
                    "SELECT COUNT(*) AS nb"
                    + base_from_f
                    + " WHERE COALESCE(td.is_suspended, 0) = 1",
                    dparams,
                )
                lignes_suspendues = int((cur.fetchone() or [0])[0] or 0)

                # DISA traitées / non traitées — hors suspendues, filtrées via JOIN
                cur.execute(
                    "SELECT COALESCE(td.statut, 'NON TRAITÉ') AS s, COUNT(*) AS nb"
                    + base_from_f
                    + " WHERE COALESCE(td.is_suspended, 0) = 0"
                    + " GROUP BY COALESCE(td.statut, 'NON TRAITÉ')",
                    dparams,
                )
                for statut, nb in cur.fetchall():
                    statut_u = str(statut or "").upper()
                    nb_i = int(nb or 0)
                    if "NON" in statut_u and "TRAIT" in statut_u:
                        lignes_non_traitees += nb_i
                    else:
                        lignes_traitees += nb_i

                # DISA par localité — filtre dans le JOIN (tous les employeurs restent)
                cur.execute(
                    f"""
                    SELECT
                        COALESCE(ie.localites, 'NON RENSEIGNÉE') AS localite,
                        SUM(CASE WHEN COALESCE(td.statut,'NON TRAITÉ')='NON TRAITÉ'
                                 THEN 1 ELSE 0 END),
                        SUM(CASE WHEN COALESCE(td.statut,'NON TRAITÉ')<>'NON TRAITÉ'
                                 THEN 1 ELSE 0 END)
                    FROM identification_employeurs ie
                    LEFT JOIN traitement_disa td ON td.employeur_id = ie.id {join_cond}
                    GROUP BY COALESCE(ie.localites,'NON RENSEIGNÉE')
                    ORDER BY 2 DESC, 3 DESC
                    """,
                    dparams,
                )
                for loc, restant, traite in cur.fetchall():
                    localites.append(str(loc))
                    disa_restantes.append(int(restant or 0))
                    disa_traitees_loc.append(int(traite or 0))

                # Secteurs — filtre dans le JOIN (tous les employeurs restent)
                cur.execute(
                    f"""
                    SELECT COALESCE(ie.secteur_activite,'NON RENSEIGNÉ') AS secteur,
                           COUNT(DISTINCT ie.id) AS nb
                    FROM identification_employeurs ie
                    LEFT JOIN traitement_disa td ON td.employeur_id = ie.id {join_cond}
                    GROUP BY COALESCE(ie.secteur_activite,'NON RENSEIGNÉ')
                    ORDER BY nb DESC LIMIT 8
                    """,
                    dparams,
                )
                for sect, nb in cur.fetchall():
                    secteurs.append(str(sect))
                    secteurs_nb.append(int(nb or 0))

                # traite_par — filtré via AND dans WHERE
                cur.execute(
                    f"""
                    SELECT COALESCE(traite_par, 'Non renseigné') AS utilisateur,
                           COUNT(*) AS nb
                    FROM traitement_disa td
                    WHERE COALESCE(statut, 'NON TRAITÉ') <> 'NON TRAITÉ'
                      {and_cond}
                    GROUP BY COALESCE(traite_par, 'Non renseigné')
                    ORDER BY nb DESC
                    """,
                    dparams,
                )
                for user, nb in cur.fetchall():
                    traite_par_users.append(str(user))
                    traite_par_nb.append(int(nb or 0))

        except Exception:
            logger.exception("Erreur lors du chargement des données du dashboard")

        # ── Snapshot des données pour l'export Excel ──────────────────
        total_disa = lignes_traitees + lignes_non_traitees + lignes_suspendues
        taux = int(lignes_traitees / total_disa * 100) if total_disa > 0 else 0
        self._last_data = {
            "date_from": self._filter_date_from,
            "date_to": self._filter_date_to,
            "total_employeurs": total_employeurs,
            "lignes_traitees": lignes_traitees,
            "lignes_non_traitees": lignes_non_traitees,
            "lignes_suspendues": lignes_suspendues,
            "taux": taux,
            "localites": localites,
            "disa_restantes": disa_restantes,
            "disa_traitees_loc": disa_traitees_loc,
            "secteurs": secteurs,
            "secteurs_nb": secteurs_nb,
            "traite_par_users": traite_par_users,
            "traite_par_nb": traite_par_nb,
        }

        # ── Grille 4 colonnes ─────────────────────────────────────────
        layout.setHorizontalSpacing(10)
        layout.setVerticalSpacing(10)
        layout.setContentsMargins(14, 12, 14, 12)
        for col in range(4):
            layout.setColumnStretch(col, 1)

        # ── LIGNE 0 : Barre de filtres ────────────────────────────────
        layout.addWidget(self._make_filter_bar(), 0, 0, 1, 4)

        # ── LIGNE 1 : 4 cartes KPI ────────────────────────────────────
        kpi_data = [
            ("Employeurs enregistrés", str(total_employeurs), "Dans la base",           _C_NAVY),
            ("DISA traitées",          str(lignes_traitees),  "Lignes validées",         _C_GREEN),
            ("DISA non traitées",      str(lignes_non_traitees), "Restantes à traiter",  _C_AMBER),
            ("Taux de traitement",     f"{taux} %",           "DISA validées / total",   _C_BLUE),
        ]
        for col, (title, value, sub, accent) in enumerate(kpi_data):
            layout.addWidget(
                self._make_kpi_card(title, value, sub, accent, delay_ms=col * 80),
                1, col,
            )

        # ── LIGNE 2 : Camembert statut | Camembert par utilisateur ────
        pie_card, pie_card_lay = self._make_chart_card(
            "Statut global des DISA", "Traitées / Non traitées / Suspendues", delay_ms=100,
        )
        pie_series = QPieSeries()
        _pie_slices = [
            (f"Traitées  ({lignes_traitees})",     lignes_traitees,    _C_BLUE,   True),
            (f"Non traitées  ({lignes_non_traitees})", lignes_non_traitees, _C_AMBER,  False),
            (f"Suspendues  ({lignes_suspendues})",  lignes_suspendues,  "#64748b", False),
        ]
        first_nonzero = True
        for label, val, color, _ in _pie_slices:
            if val <= 0:
                continue
            sl = pie_series.append(label, val)
            sl.setColor(QColor(color)); sl.setBorderColor(QColor(color))
            if first_nonzero:
                sl.setExploded(True); sl.setExplodeDistanceFactor(0.05)
                first_nonzero = False
        if pie_series.count() == 0:
            sl = pie_series.append("Aucune donnée", 1)
            sl.setColor(QColor(_BORDER)); sl.setBorderColor(QColor(_BORDER))
        pie_chart = self._base_chart()
        pie_chart.addSeries(pie_series)
        pie_chart.setMargins(QMargins(8, 4, 8, 4))
        self._style_legend(pie_chart)
        pie_card_lay.addWidget(self._styled_chart_view(pie_chart, min_h=200))
        layout.addWidget(pie_card, 2, 0, 1, 2)

        user_pie_card, user_pie_card_lay = self._make_chart_card(
            "DISA traitées par utilisateur",
            "Contribution de chaque agent au traitement",
            delay_ms=180,
        )
        user_pie_series = QPieSeries()
        if traite_par_users:
            for i, (user, nb) in enumerate(zip(traite_par_users, traite_par_nb)):
                sl = user_pie_series.append(f"{user}  ({nb})", nb)
                c = _USER_COLORS[i % len(_USER_COLORS)]
                sl.setColor(QColor(c)); sl.setBorderColor(QColor(c))
                if i == 0:
                    sl.setExploded(True); sl.setExplodeDistanceFactor(0.05)
        else:
            sl = user_pie_series.append("Aucune donnée", 1)
            sl.setColor(QColor(_BORDER)); sl.setBorderColor(QColor(_BORDER))
        user_pie_chart = self._base_chart()
        user_pie_chart.addSeries(user_pie_series)
        user_pie_chart.setMargins(QMargins(8, 4, 8, 4))
        self._style_legend(user_pie_chart)
        user_pie_card_lay.addWidget(self._styled_chart_view(user_pie_chart, min_h=200))
        layout.addWidget(user_pie_card, 2, 2, 1, 2)

        # ── LIGNE 3 : Histogramme DISA par localité ───────────────────
        bar_card, bar_card_lay = self._make_chart_card(
            "DISA restantes et traitées par localité",
            "Comparaison du volume de traitement par localité d'employeur",
            delay_ms=260,
        )
        chart_bar = self._base_chart()
        chart_bar.setMargins(QMargins(12, 8, 12, 8))
        series_bar = QBarSeries()
        series_bar.setBarWidth(0.65)
        series_bar.setLabelsVisible(True)

        set_restantes = QBarSet("DISA restantes")
        set_restantes.setColor(QColor(_C_AMBER)); set_restantes.setBorderColor(QColor(_C_AMBER))
        set_traitees_bar = QBarSet("DISA traitées")
        set_traitees_bar.setColor(QColor(_C_BLUE)); set_traitees_bar.setBorderColor(QColor(_C_BLUE))

        if localites:
            for r in disa_restantes: set_restantes.append(r)
            for t in disa_traitees_loc: set_traitees_bar.append(t)
        else:
            localites = ["Aucune donnée"]
            set_restantes.append(0); set_traitees_bar.append(0)

        series_bar.append(set_restantes); series_bar.append(set_traitees_bar)
        chart_bar.addSeries(series_bar)

        ax_x_bar = QBarCategoryAxis()
        ax_x_bar.append([_truncate(l, 18) for l in localites])
        ax_x_bar.setLabelsAngle(-30)
        self._style_axis(ax_x_bar)
        chart_bar.addAxis(ax_x_bar, Qt.AlignBottom)
        series_bar.attachAxis(ax_x_bar)

        max_y = max([max(disa_restantes or [0]), max(disa_traitees_loc or [0])])
        ax_y_bar = QValueAxis()
        ax_y_bar.setRange(0, max_y * 1.2 if max_y > 0 else 1)
        self._style_axis(ax_y_bar)
        chart_bar.addAxis(ax_y_bar, Qt.AlignLeft)
        series_bar.attachAxis(ax_y_bar)
        self._style_legend(chart_bar)

        def show_detail(status: bool, index: int, _serie: str) -> None:
            if not status or index < 0 or index >= len(localites):
                if self._detail_dialog is not None: self._detail_dialog.hide()
                return
            loc_name = localites[index]
            r_val = disa_restantes[index]; t_val = disa_traitees_loc[index]
            if self._detail_dialog is None:
                self._detail_dialog = QDialog()
                self._detail_dialog.setWindowTitle("Détail DISA par localité")
                self._detail_dialog.setModal(False)
                vb = QVBoxLayout(self._detail_dialog)
                self._detail_label = QLabel()
                self._detail_label.setStyleSheet("font-size: 13px; font-weight: 600; padding: 8px;")
                vb.addWidget(self._detail_label)
            if self._detail_label is not None:
                self._detail_label.setText(
                    f"<b>Localité :</b> {loc_name}<br>"
                    f"<b>DISA restantes :</b> {r_val}<br>"
                    f"<b>DISA traitées :</b> {t_val}"
                )
            self._detail_dialog.setWindowFlags(Qt.ToolTip)
            self._detail_dialog.adjustSize()
            pos = QCursor.pos()
            self._detail_dialog.move(pos.x() + 15, pos.y() + 15)
            self._detail_dialog.show()

        try:
            set_restantes.hovered.connect(lambda s, i: show_detail(s, i, "restantes"))
            set_traitees_bar.hovered.connect(lambda s, i: show_detail(s, i, "traitees"))
        except Exception:
            pass

        view_bar = self._styled_chart_view(chart_bar, min_h=320)
        view_bar.setMaximumHeight(480)
        bar_card_lay.addWidget(view_bar)
        layout.addWidget(bar_card, 3, 0, 1, 4)

        # ── LIGNE 4 : Barres horizontales par secteur ─────────────────
        sect_card, sect_card_lay = self._make_chart_card(
            "Employeurs par secteur d'activité",
            "Top secteurs — nombre d'employeurs enregistrés",
            delay_ms=340,
        )
        if not secteurs:
            secteurs = ["Aucune donnée"]; secteurs_nb = [0]

        secteurs_display = [_truncate(s, 24) for s in secteurs]
        sect_set = QBarSet("Employeurs")
        sect_set.setColor(QColor(_C_NAVY)); sect_set.setBorderColor(QColor(_C_NAVY))
        for n in secteurs_nb: sect_set.append(n)

        sect_series = QHorizontalBarSeries()
        sect_series.append(sect_set)
        sect_series.setBarWidth(0.55)
        sect_series.setLabelsVisible(True)

        sect_chart = self._base_chart()
        sect_chart.addSeries(sect_series)
        sect_chart.setMargins(QMargins(8, 4, 24, 4))

        ax_y_sect = QBarCategoryAxis()
        ax_y_sect.append(secteurs_display)
        self._style_axis(ax_y_sect)
        sect_chart.addAxis(ax_y_sect, Qt.AlignLeft)
        sect_series.attachAxis(ax_y_sect)

        max_sect = max(secteurs_nb) if secteurs_nb else 1
        ax_x_sect = QValueAxis()
        ax_x_sect.setRange(0, max_sect * 1.3)
        ax_x_sect.setLabelFormat("%d")
        ax_x_sect.setTickCount(5)
        self._style_axis(ax_x_sect)
        sect_chart.addAxis(ax_x_sect, Qt.AlignBottom)
        sect_series.attachAxis(ax_x_sect)
        sect_chart.legend().setVisible(False)

        min_h_sect = max(200, len(secteurs) * 32 + 60)
        sect_card_lay.addWidget(self._styled_chart_view(sect_chart, min_h=min_h_sect))
        layout.addWidget(sect_card, 4, 0, 1, 4)

    # ── Export Excel ─────────────────────────────────────────────────────────

    def _export_to_excel(self) -> None:
        """Génère un rapport Excel multi-feuilles à partir des données du dashboard."""
        from datetime import datetime
        from PySide6.QtWidgets import QFileDialog, QMessageBox

        if not self._last_data:
            QMessageBox.warning(None, "Export", "Aucune donnée à exporter. Ouvrez d'abord le tableau de bord.")
            return

        default_name = f"Rapport_DiSA_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            None,
            "Enregistrer le rapport Excel",
            default_name,
            "Fichier Excel (*.xlsx)",
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(None, "Dépendance manquante", "openpyxl n'est pas installé.\nExécutez : pip install openpyxl")
            return

        d = self._last_data
        now_str = datetime.now().strftime("%d/%m/%Y à %H:%M")
        period = f"Du {d['date_from']} au {d['date_to']}"

        # ── Styles communs ──────────────────────────────────────────────────
        NAVY    = "1E3A5F"
        NAVY2   = "2A4F80"
        GREEN   = "15803D"
        AMBER   = "B45309"
        BLUE    = "1D4ED8"
        WHITE   = "FFFFFF"
        LIGHT   = "F0F4FB"
        GREY    = "F3F4F6"
        BORDER_C = "CBD5E1"

        def _fill(hex_color: str) -> PatternFill:
            return PatternFill("solid", fgColor=hex_color)

        def _font(bold=False, color=None, size=11) -> Font:
            return Font(bold=bold, color=color or "111827", size=size)

        def _border() -> Border:
            thin = Side(style="thin", color=BORDER_C)
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        def _center(wrap=False) -> Alignment:
            return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

        def _left(wrap=False) -> Alignment:
            return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

        def _apply_header_row(ws, row: int, values: list[str], widths: list[int]) -> None:
            """Écrit une ligne d'en-tête bleue."""
            for col, (val, w) in enumerate(zip(values, widths), start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = _fill(NAVY)
                cell.font = _font(bold=True, color=WHITE, size=10)
                cell.alignment = _center()
                cell.border = _border()
                ws.column_dimensions[get_column_letter(col)].width = w

        def _apply_data_row(ws, row: int, values: list, even: bool, formats=None) -> None:
            """Écrit une ligne de données alternée blanc/gris clair."""
            bg = LIGHT if even else WHITE
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = _fill(bg)
                cell.font = _font(size=10)
                cell.alignment = _center() if isinstance(val, (int, float)) else _left()
                cell.border = _border()
                if formats and col <= len(formats) and formats[col - 1]:
                    cell.number_format = formats[col - 1]

        def _title_block(ws, title: str, subtitle: str, period_str: str, generated: str) -> None:
            """Bloc titre en haut de chaque feuille."""
            ws.merge_cells("A1:F1")
            t = ws["A1"]
            t.value = f"IPS-CNPS — Agence de Gagnoa  |  {title}"
            t.fill = _fill(NAVY)
            t.font = _font(bold=True, color=WHITE, size=14)
            t.alignment = _center()

            ws.merge_cells("A2:F2")
            s = ws["A2"]
            s.value = subtitle
            s.fill = _fill(NAVY2)
            s.font = _font(color=WHITE, size=10)
            s.alignment = _center()

            ws.merge_cells("A3:C3")
            ws["A3"].value = f"Période : {period_str}"
            ws["A3"].fill = _fill(GREY)
            ws["A3"].font = _font(bold=True, size=10)
            ws["A3"].alignment = _left()

            ws.merge_cells("D3:F3")
            ws["D3"].value = f"Généré le : {generated}"
            ws["D3"].fill = _fill(GREY)
            ws["D3"].font = _font(size=10)
            ws["D3"].alignment = _center()

            ws.row_dimensions[1].height = 26
            ws.row_dimensions[2].height = 18
            ws.row_dimensions[3].height = 16

        wb = openpyxl.Workbook()

        # ══════════════════════════════════════════════════════════════════════
        # FEUILLE 1 — Synthèse
        # ══════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Synthèse"
        ws1.sheet_view.showGridLines = False

        _title_block(ws1, "Tableau de Bord DiSA", "Synthèse des indicateurs clés", period, now_str)

        # KPI table
        ws1.row_dimensions[5].height = 14
        _apply_header_row(ws1, 6, ["Indicateur", "Valeur", "Description"], [32, 18, 40])

        kpi_rows = [
            ("Employeurs enregistrés",  d["total_employeurs"],    "Nombre total d'employeurs dans la base"),
            ("DISA traitées",           d["lignes_traitees"],     "Dossiers validés sur la période"),
            ("DISA non traitées",       d["lignes_non_traitees"], "Dossiers restants à traiter"),
            ("Taux de traitement",      f"{d['taux']} %",         "Part des DISA traitées / total"),
        ]
        for i, (ind, val, desc) in enumerate(kpi_rows, start=7):
            even = (i % 2 == 0)
            _apply_data_row(ws1, i, [ind, val, desc], even)
            # Coloriser la valeur selon l'indicateur
            val_cell = ws1.cell(row=i, column=2)
            if "traitées" in ind.lower() and "non" not in ind.lower():
                val_cell.font = _font(bold=True, color=GREEN, size=11)
            elif "non" in ind.lower():
                val_cell.font = _font(bold=True, color=AMBER, size=11)
            elif "taux" in ind.lower():
                val_cell.font = _font(bold=True, color=BLUE, size=11)
            else:
                val_cell.font = _font(bold=True, color=NAVY, size=11)
            val_cell.alignment = _center()

        # Note de bas de page
        note_row = 7 + len(kpi_rows) + 2
        ws1.merge_cells(f"A{note_row}:F{note_row}")
        note = ws1[f"A{note_row}"]
        note.value = "Rapport généré automatiquement par DisaManager — CNPS Gagnoa"
        note.font = Font(italic=True, color="6B7280", size=9)
        note.alignment = _left()

        # ══════════════════════════════════════════════════════════════════════
        # FEUILLE 2 — Par localité
        # ══════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Par localité")
        ws2.sheet_view.showGridLines = False
        _title_block(ws2, "DISA par Localité", "Répartition des dossiers par localité d'employeur", period, now_str)

        ws2.row_dimensions[5].height = 14
        _apply_header_row(ws2, 6,
            ["Localité", "DISA restantes", "DISA traitées", "Total", "Taux (%)"],
            [28, 18, 18, 14, 14],
        )
        locs = d["localites"] or ["Aucune donnée"]
        reste = d["disa_restantes"] or [0]
        traite = d["disa_traitees_loc"] or [0]
        for i, (loc, r, t) in enumerate(zip(locs, reste, traite), start=7):
            total = r + t
            tx = f"{int(t / total * 100)} %" if total > 0 else "—"
            _apply_data_row(ws2, i, [loc, r, t, total, tx], i % 2 == 0)
            # Coloriser DISA traitées en vert
            ws2.cell(row=i, column=3).font = _font(bold=True, color=GREEN, size=10)

        # Ligne total
        total_row = 7 + len(locs)
        ws2.cell(row=total_row, column=1).value = "TOTAL"
        ws2.cell(row=total_row, column=1).font = _font(bold=True, color=WHITE, size=10)
        ws2.cell(row=total_row, column=1).fill = _fill(NAVY)
        ws2.cell(row=total_row, column=1).alignment = _center()
        ws2.cell(row=total_row, column=2).value = sum(reste)
        ws2.cell(row=total_row, column=2).fill = _fill(NAVY); ws2.cell(row=total_row, column=2).font = _font(bold=True, color=WHITE, size=10); ws2.cell(row=total_row, column=2).alignment = _center()
        ws2.cell(row=total_row, column=3).value = sum(traite)
        ws2.cell(row=total_row, column=3).fill = _fill(NAVY); ws2.cell(row=total_row, column=3).font = _font(bold=True, color=WHITE, size=10); ws2.cell(row=total_row, column=3).alignment = _center()
        ws2.cell(row=total_row, column=4).value = sum(reste) + sum(traite)
        ws2.cell(row=total_row, column=4).fill = _fill(NAVY); ws2.cell(row=total_row, column=4).font = _font(bold=True, color=WHITE, size=10); ws2.cell(row=total_row, column=4).alignment = _center()
        grand_total = sum(reste) + sum(traite)
        grand_taux = f"{int(sum(traite) / grand_total * 100)} %" if grand_total > 0 else "—"
        ws2.cell(row=total_row, column=5).value = grand_taux
        ws2.cell(row=total_row, column=5).fill = _fill(NAVY); ws2.cell(row=total_row, column=5).font = _font(bold=True, color=WHITE, size=10); ws2.cell(row=total_row, column=5).alignment = _center()

        # ══════════════════════════════════════════════════════════════════════
        # FEUILLE 3 — Par secteur
        # ══════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Par secteur")
        ws3.sheet_view.showGridLines = False
        _title_block(ws3, "Employeurs par Secteur", "Répartition des employeurs par secteur d'activité", period, now_str)

        ws3.row_dimensions[5].height = 14
        _apply_header_row(ws3, 6, ["Secteur d'activité", "Nb employeurs", "Part (%)"], [36, 16, 14])
        sects = d["secteurs"] or ["Aucune donnée"]
        sects_nb = d["secteurs_nb"] or [0]
        total_sect = sum(sects_nb) or 1
        for i, (sect, nb) in enumerate(zip(sects, sects_nb), start=7):
            part = f"{round(nb / total_sect * 100, 1)} %"
            _apply_data_row(ws3, i, [sect, nb, part], i % 2 == 0)

        # ══════════════════════════════════════════════════════════════════════
        # FEUILLE 4 — Par agent
        # ══════════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet("Par agent")
        ws4.sheet_view.showGridLines = False
        _title_block(ws4, "Traitements par Agent", "Nombre de DISA traitées par agent sur la période", period, now_str)

        ws4.row_dimensions[5].height = 14
        _apply_header_row(ws4, 6, ["Agent", "DISA traitées", "Part (%)"], [30, 16, 14])
        users = d["traite_par_users"] or []
        nbs   = d["traite_par_nb"] or []
        total_agents = sum(nbs) or 1
        for i, (user, nb) in enumerate(zip(users, nbs), start=7):
            part = f"{round(nb / total_agents * 100, 1)} %"
            _apply_data_row(ws4, i, [user, nb, part], i % 2 == 0)

        if not users:
            ws4.cell(row=7, column=1).value = "Aucune donnée disponible"
            ws4.cell(row=7, column=1).font = Font(italic=True, color="6B7280", size=10)

        # ── Enregistrement ──────────────────────────────────────────────────
        try:
            wb.save(path)
        except Exception as exc:
            QMessageBox.critical(None, "Erreur d'enregistrement", f"Impossible d'écrire le fichier :\n{exc}")
            return

        reply = QMessageBox.information(
            None,
            "Export réussi",
            f"Rapport Excel généré :\n{path}\n\nVoulez-vous l'ouvrir ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if reply == QMessageBox.StandardButton.Yes:
            import sys, subprocess, os
            if sys.platform == "darwin":
                subprocess.run(["open", path], check=False)
            elif sys.platform == "win32":
                os.startfile(path)  # type: ignore[attr-defined]
            else:
                subprocess.run(["xdg-open", path], check=False)

    # ── Responsive fonts ──────────────────────────────────────────────────────

    def update_font_sizes(self, scale: float) -> None:
        def _clamp(size: int, lo: int = 8, hi: int = 26) -> int:
            return max(lo, min(hi, size))
        for lbl, base_size in self._responsive_label_specs:
            new_size = _clamp(int(base_size * scale))
            font = QFont(lbl.font())
            font.setPointSize(new_size)
            lbl.setFont(font)
